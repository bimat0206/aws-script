#!/usr/bin/env python3
"""
AWS Cost Calculator JSON to Excel Exporter
This script parses an AWS Cost Calculator JSON export and populates a formatted Excel template.
"""

import json
import re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from collections import defaultdict
import glob
import os
import sys

# AWS instance specifications database (vCPU and Memory)
# Format: instance_type: (vCPU, Memory_GB)
INSTANCE_SPECS = {
    # T-series (Burstable)
    "t3.medium": (2, 4), "t3.large": (2, 8), "t3.xlarge": (4, 16), "t3.2xlarge": (8, 32),
    "t4g.medium": (2, 4), "t4g.large": (2, 8), "t4g.xlarge": (4, 16), "t4g.2xlarge": (8, 32),
    
    # M-series (General Purpose)
    "m7g.xlarge": (4, 16), "m7g.2xlarge": (8, 32), "m7g.4xlarge": (16, 64), "m7g.8xlarge": (32, 128),
    "m7i.xlarge": (4, 16), "m7i.2xlarge": (8, 32), "m7i.4xlarge": (16, 64), "m7i.8xlarge": (32, 128),
    "m7i-flex.large": (2, 8), "m7i-flex.xlarge": (4, 16), "m7i-flex.2xlarge": (8, 32),
    
    # C-series (Compute Optimized)
    "c7g.xlarge": (4, 8), "c7g.2xlarge": (8, 16), "c7g.4xlarge": (16, 32), "c7g.8xlarge": (32, 64),
    "c7i.4xlarge": (16, 32), "c7i.8xlarge": (32, 64),
    "c7i-flex.xlarge": (4, 8), "c7i-flex.2xlarge": (8, 16),
    
    # R-series (Memory Optimized)
    "r7g.xlarge": (4, 32), "r7g.2xlarge": (8, 64), "r7g.4xlarge": (16, 128), "r7g.8xlarge": (32, 256),
    "r7i.large": (2, 16), "r7i.xlarge": (4, 32), "r7i.2xlarge": (8, 64), "r7i.4xlarge": (16, 128),
}

def get_instance_specs(instance_type):
    """
    Extract vCPU and Memory from instance type.
    Returns (vCPU, Memory_GB) tuple or ("N/A", "N/A") if not found.
    """
    if not instance_type or instance_type == "-":
        return ("N/A", "N/A")
    
    # Clean instance type string
    instance_type = instance_type.strip().lower()
    
    # Look up in database
    if instance_type in INSTANCE_SPECS:
        return INSTANCE_SPECS[instance_type]
    
    # If not found, try to parse from common patterns
    # Example: c7g.4xlarge -> 16 vCPU, 32 GB (Compute optimized)
    return ("N/A", "N/A")

def extract_instance_type(properties):
    """Extract instance type from Properties dict."""
    if not properties:
        return "-"
    
    # Look for instance type in various possible keys
    for key in ["Advance EC2 instance", "Instance type", "Instance Type"]:
        if key in properties:
            return properties[key]
    
    return "-"

def extract_quantity(properties):
    """Extract quantity from Properties dict."""
    if not properties:
        return 1
    
    # Look for quantity in Workload field
    workload = properties.get("Workload", "")
    if "Number of instances:" in workload:
        match = re.search(r'Number of instances:\s*(\d+)', workload)
        if match:
            return int(match.group(1))
    
    # Look for other quantity fields
    for key in ["Number of instances", "Quantity"]:
        if key in properties:
            try:
                return int(properties[key])
            except (ValueError, TypeError):
                pass
    
    return 1

def extract_storage(properties):
    """Extract EBS storage amount from Properties dict."""
    if not properties:
        return 0
    
    storage = properties.get("EBS Storage amount", "0 GB")
    match = re.search(r'(\d+)', str(storage))
    if match:
        return int(match.group(1))
    
    return 0

def extract_os(service_name, properties):
    """Extract Operating System / DBMS information."""
    if not properties:
        return "AWS Managed"
    
    # For EC2, return the operating system
    if "EC2" in service_name:
        return properties.get("Operating system", "Linux")
    
    # For other services, return "AWS Managed"
    return "AWS Managed"

def clean_service_properties(properties, service_name):
    """
    Clean service properties by removing 'Tenancy' and formatting nicely.
    """
    if not properties:
        return ""
    
    # Keys to exclude
    exclude_keys = {"Tenancy"}
    
    # Build property string
    prop_parts = []
    for key, value in properties.items():
        if key not in exclude_keys and value:
            # Clean up the value
            value_str = str(value).strip()
            if value_str and value_str != "0":
                prop_parts.append(f"{key}: {value_str}")
    
    return "; ".join(prop_parts)

def parse_json_data(json_file_path):
    """
    Parse AWS Cost Calculator JSON and extract relevant data.
    Returns list of dicts with structured data.
    """
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    rows = []
    
    # Navigate through nested structure: Groups -> environments -> categories -> Services
    groups = data.get("Groups", {})
    
    for env_name, env_data in groups.items():
        # Environment level (staging, ops, prod, etc.)
        
        for category_name, category_data in env_data.items():
            # Category level (compute, network, db, storage, etc.)
            group_name = category_name.strip()
            
            # Handle nested structure (some categories have sub-categories)
            if "Services" in category_data:
                services = category_data.get("Services", [])
            else:
                # Nested structure like prod -> compute -> aggregator-cluster -> Services
                services = []
                for sub_category_name, sub_category_data in category_data.items():
                    if isinstance(sub_category_data, dict) and "Services" in sub_category_data:
                        services.extend(sub_category_data.get("Services", []))
            
            # Process each service
            for service in services:
                service_name = service.get("Service Name", "")
                description = service.get("Description", "N/A")
                region = service.get("Region", "N/A")
                properties = service.get("Properties", {})
                
                # Extract cost
                service_cost = service.get("Service Cost", {})
                monthly_cost = service_cost.get("monthly", "0.00")
                try:
                    total_cost = float(monthly_cost)
                except (ValueError, TypeError):
                    total_cost = 0.0
                
                # Extract instance details
                instance_type = extract_instance_type(properties)
                vcpu, memory = get_instance_specs(instance_type)
                storage = extract_storage(properties)
                quantity = extract_quantity(properties)
                os_dbms = extract_os(service_name, properties)
                service_props = clean_service_properties(properties, service_name)
                
                # Build row
                row = {
                    "Region": region,
                    "Environment": env_name,
                    "Group": group_name,  # Add Group
                    "Service Name": service_name.strip(),
                    "Instance Type": instance_type if "EC2" in service_name else "-",
                    "vCPU (Core)": vcpu,
                    "Memory (GB)": memory,
                    "Storage (GB)": storage,
                    "Resources Name": description,
                    "Service Properties": service_props,
                    "Quantity": quantity,
                    "OS/DBMS": os_dbms,
                    "Total Cost": total_cost
                }
                
                rows.append(row)
    
    return rows

def apply_cell_borders(ws, start_row, end_row, start_col, end_col):
    """Apply thin black borders to all cells in the specified range."""
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

def merge_cells_by_column(ws, start_row, end_row, col_idx, data_rows):
    """
    Merge cells vertically for consecutive rows with same value.
    
    Args:
        ws: worksheet object
        start_row: starting row number
        end_row: ending row number
        col_idx: column index (1-based)
        data_rows: list of data dictionaries
    """
    if not data_rows:
        return
    
    # Column name mapping (adjusted for template layout)
    col_names = {
        2: "Environment",  # Column B
        3: "Group",        # Column C
        6: "Region"        # Column F
    }
    
    if col_idx not in col_names:
        return
    
    col_name = col_names[col_idx]
    
    # Find groups of consecutive rows with same value
    i = 0
    while i < len(data_rows):
        value = data_rows[i][col_name]
        group_start = i
        group_end = i
        
        # Find end of group
        while group_end + 1 < len(data_rows) and data_rows[group_end + 1][col_name] == value:
            group_end += 1
        
        # Merge if group has more than 1 cell
        if group_end > group_start:
            merge_start_row = start_row + group_start
            merge_end_row = start_row + group_end
            ws.merge_cells(start_row=merge_start_row, start_column=col_idx, 
                          end_row=merge_end_row, end_column=col_idx)
            
            # Center align merged cells
            top_cell = ws.cell(row=merge_start_row, column=col_idx)
            top_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        i = group_end + 1

def auto_fit_columns(ws, min_width=10, max_width=100):
    """
    Adjust column widths to fit content.
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass
        
        # Calculate adjusted width (add buffer for padding)
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def auto_adjust_row_heights(ws, start_row, end_row, default_height=30):
    """
    Automatically adjust row heights based on wrapped text content.
    """
    for row_num in range(start_row, end_row + 1):
        # Set a reasonable default height for data rows to accommodate wrapped text
        ws.row_dimensions[row_num].height = default_height

def setup_header_row(ws, header_row=11, start_col=2):
    """
    Create and format header row with column names.
    Template has headers at row 11, starting from column B (2).
    """
    # Column order matching template: B=Environment, C=Group, D=Resources Name, 
    # E=Service Name, F=Region, G=Instance type, H=Services Properties, I=OS/DBMS, 
    # J=vCPU, K=Memory, L=Storage, M=Quantity, N=Service Cost
    headers = [
        "Environment",
        "Group",
        "Resources Name",
        "Service Name",
        "Region",
        "Instance Type",
        "Service Properties",
        "OS/DBMS",
        "vCPU\n(Core)",
        "Memory\n(GB)",
        "Storage\n(GB)",
        "Quantity",
        "Service Cost (monthly)"
    ]
    
    # Header styling
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Apply headers
    for idx, header_text in enumerate(headers):
        col_idx = start_col + idx
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Set row height for header
    ws.row_dimensions[header_row].height = 30

def apply_data_cell_formatting(ws, row_num, col_idx, value):
    """
    Apply formatting to data cells.
    """
    cell = ws.cell(row=row_num, column=col_idx)
    cell.value = value
    
    # Center align for specific columns (based on template layout)
    # Cols: B=Env, C=Grp, D=Resources, E=Service, F=Region, G=Instance, H=Props, I=OS, J=vCPU, K=Mem, L=Storage, M=Qty, N=Cost
    if col_idx in [6, 7, 9, 10, 11, 12, 13]:  # Region, Instance Type, OS/DBMS, vCPU, Memory, Storage, Quantity
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    return cell

def populate_excel(json_file, excel_template, output_file=None):
    """
    Main function to populate Excel template with JSON data.
    
    Args:
        json_file: path to AWS Cost Calculator JSON export
        excel_template: path to Excel template file
        output_file: path for output file (if None, overwrites template)
    """
    if output_file is None:
        output_file = excel_template.replace('.xlsx', '_populated.xlsx')
    
    print(f"Reading JSON data from: {json_file}")
    data_rows = parse_json_data(json_file)
    print(f"Parsed {len(data_rows)} service items from JSON")
    
    print(f"Loading Excel template: {excel_template}")
    wb = load_workbook(excel_template)
    ws = wb.active
    
    # Unmerge cells in data area (rows 11+) while preserving template branding (rows 1-10)
    print("Unmerging data area cells...")
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        # Only unmerge if the range starts at row 11 or later (data area)
        if merged_range.min_row >= 11:
            ws.unmerge_cells(str(merged_range))
    
    # Setup header row at row 11, starting from column B (2)
    print("Setting up header row...")
    setup_header_row(ws, header_row=11, start_col=2)
    
    # Data starts at row 12 (after header at row 11), column B (2)
    start_row = 12
    start_col = 2
    print(f"Writing data starting at row {start_row}, column {start_col}")
    
    # Write data with formatting
    # Template column mapping: B=Environment, C=Group, D=Resources Name, E=Service Name, F=Region,
    # G=Instance type, H=Services Properties, I=OS/DBMS, J=vCPU, K=Memory, L=Storage, M=Quantity, N=Cost
    
    total_monthly_cost = 0.0
    
    for idx, row_data in enumerate(data_rows):
        row_num = start_row + idx
        
        # Determine if service has compute specs (EC2 or RDS)
        service_name = row_data["Service Name"]
        has_compute_specs = "EC2" in service_name or "RDS" in service_name
        
        # Column mapping with formatting (matching template layout)
        apply_data_cell_formatting(ws, row_num, 2, row_data["Environment"])  # B
        apply_data_cell_formatting(ws, row_num, 3, row_data["Group"])   # C
        apply_data_cell_formatting(ws, row_num, 4, row_data["Resources Name"])  # D
        apply_data_cell_formatting(ws, row_num, 5, service_name)  # E
        apply_data_cell_formatting(ws, row_num, 6, row_data["Region"])  # F
        apply_data_cell_formatting(ws, row_num, 7, row_data["Instance Type"])  # G
        apply_data_cell_formatting(ws, row_num, 8, row_data["Service Properties"])  # H
        apply_data_cell_formatting(ws, row_num, 9, row_data["OS/DBMS"])  # I
        
        # Only show vCPU, Memory, Storage for EC2/RDS services
        apply_data_cell_formatting(ws, row_num, 10, row_data["vCPU (Core)"] if has_compute_specs else "-")  # J
        apply_data_cell_formatting(ws, row_num, 11, row_data["Memory (GB)"] if has_compute_specs else "-")  # K
        apply_data_cell_formatting(ws, row_num, 12, row_data["Storage (GB)"] if has_compute_specs else "-")  # L
        apply_data_cell_formatting(ws, row_num, 13, row_data["Quantity"])  # M
        
        # Format Total Cost as currency (Column N)
        cost_cell = apply_data_cell_formatting(ws, row_num, 14, row_data["Total Cost"])
        cost_cell.number_format = '$#,##0.00'
        cost_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Accumulate total cost
        total_monthly_cost += row_data["Total Cost"]
    
    end_row = start_row + len(data_rows) - 1
    
    # Apply borders to data range (columns B to N = 2 to 14)
    print("Applying cell borders...")
    apply_cell_borders(ws, start_row, end_row, 2, 14)
    
    # Merge cells for Environment and Region and Group columns
    print("Merging cells for Environment and Region and Group...")
    merge_cells_by_column(ws, start_row, end_row, 2, data_rows)  # Environment (column B)
    merge_cells_by_column(ws, start_row, end_row, 3, data_rows)  # Group (column C)
    merge_cells_by_column(ws, start_row, end_row, 6, data_rows)  # Region (column F)
    
    # Auto-fit columns
    print("Auto-fitting column widths...")
    auto_fit_columns(ws)
    
    # Auto-adjust row heights for better text visibility
    print("Auto-adjusting row heights...")
    auto_adjust_row_heights(ws, start_row, end_row, default_height=40)
    
    # --- Add Total Rows ---
    print("Adding total payment rows...")
    total_start_row = end_row + 2
    
    # Styling for totals
    total_label_font = Font(bold=True)
    total_value_font = Font(bold=True)
    total_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 1. Total Monthly Payment
    row_monthly = total_start_row
    # Merge label cells (B to M -> 2 to 13)
    ws.merge_cells(start_row=row_monthly, start_column=2, end_row=row_monthly, end_column=13)
    label_cell_m = ws.cell(row=row_monthly, column=2)
    label_cell_m.value = "TOTAL MONTH PAYMENT"
    label_cell_m.font = total_label_font
    label_cell_m.alignment = Alignment(horizontal="right", vertical="center")
    
    # Apply border to merged label cells
    for col in range(2, 14):
        ws.cell(row=row_monthly, column=col).border = total_border
        
    # Value cell (N -> 14)
    val_cell_m = ws.cell(row=row_monthly, column=14)
    val_cell_m.value = total_monthly_cost
    val_cell_m.font = total_value_font
    val_cell_m.number_format = '$#,##0.00'
    val_cell_m.alignment = Alignment(horizontal="right", vertical="center")
    val_cell_m.border = total_border
    
    # 2. Total Yearly Payment
    row_yearly = total_start_row + 1
    # Merge label cells (B to M -> 2 to 13)
    ws.merge_cells(start_row=row_yearly, start_column=2, end_row=row_yearly, end_column=13)
    label_cell_y = ws.cell(row=row_yearly, column=2)
    label_cell_y.value = "TOTAL YEARLY (1 YEAR) PAYMENT"
    label_cell_y.font = total_label_font
    label_cell_y.alignment = Alignment(horizontal="right", vertical="center")
    
    # Apply border to merged label cells
    for col in range(2, 14):
        ws.cell(row=row_yearly, column=col).border = total_border

    # Value cell (N -> 14)
    val_cell_y = ws.cell(row=row_yearly, column=14)
    val_cell_y.value = total_monthly_cost * 12
    val_cell_y.font = total_value_font
    val_cell_y.number_format = '$#,##0.00'
    val_cell_y.alignment = Alignment(horizontal="right", vertical="center")
    val_cell_y.border = total_border
    
    # Adjust heights for total rows
    ws.row_dimensions[row_monthly].height = 30
    ws.row_dimensions[row_yearly].height = 30
    
    # Save workbook
    print(f"Saving output to: {output_file}")
    wb.save(output_file)
    print("✓ Excel file successfully created!")
    
    return output_file

if __name__ == "__main__":
    # Template and output settings
    template_file = "template.xlsx"
    output_filename = "AWS_Cost_Report.xlsx"
    
    # 1. Scan for JSON files in the current directory
    current_dir = os.getcwd()
    json_files = glob.glob(os.path.join(current_dir, "*.json"))
    
    if not json_files:
        print(f"❌ No JSON files found in {current_dir}")
        print("Please ensure your AWS Cost Calculator JSON export is in this folder.")
        sys.exit(1)
        
    # 2. Interactive selection
    print(f"\nFound {len(json_files)} JSON file(s):")
    for i, f in enumerate(json_files):
        print(f"[{i+1}] {os.path.basename(f)}")
        
    selected_file = None
    while selected_file is None:
        try:
            selection = input("\nSelect a file number to parse (or 'q' to quit): ").strip()
            if selection.lower() == 'q':
                print("Exiting...")
                sys.exit(0)
                
            idx = int(selection) - 1
            if 0 <= idx < len(json_files):
                selected_file = json_files[idx]
            else:
                print("Invalid selection. Please try again.")
        except ValueError:
            print("Invalid input. Please enter a number.")
            
    print(f"\nSelected: {os.path.basename(selected_file)}")
    
    # Determine output path based on input name to avoid overwriting if processing multiple
    base_name = os.path.splitext(os.path.basename(selected_file))[0]
    output_file = os.path.join(current_dir, f"{base_name}_Report_populated.xlsx")
    
    # Check if template exists
    if not os.path.exists(template_file):
        # Try finding template in the same folder as script if not in current cwd
        script_dir = os.path.dirname(os.path.abspath(__file__))
        potential_template = os.path.join(script_dir, "template.xlsx")
        if os.path.exists(potential_template):
            template_file = potential_template
        else:
            print(f"⚠️ Template file '{template_file}' not found.")
            # Depending on logic, might want to exit or let openpyxl fail
    
    try:
        populate_excel(selected_file, template_file, output_file)
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
