#!/usr/bin/env python3
"""
Export AWS VPC Route Tables and Transit Gateway Route Tables to separate Excel sheets.

Features:
- Interactive AWS profile and region selection
- Exports VPC Route Tables with detailed routing information
- Exports Transit Gateway Route Tables with propagated and static routes
- Generates .xlsx with structured, formatted tables in separate sheets

Requirements:
    pip install boto3 pandas openpyxl
Optional (faster Excel writing):
    pip install xlsxwriter

Usage:
    python export_route_tables_to_xlsx.py
"""

import sys
import os
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from collections import defaultdict

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError, ProfileNotFound
except ImportError:
    print("Missing dependency: boto3. Install with `pip install boto3`.", file=sys.stderr)
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("Missing dependency: pandas. Install with `pip install pandas`.", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Missing dependency: openpyxl. Install with `pip install openpyxl`.", file=sys.stderr)
    sys.exit(1)


def tag_name(tags: Optional[List[Dict]]) -> str:
    """Extract the Name tag value from a list of tags."""
    if not tags:
        return ""
    for t in tags:
        if t.get("Key") == "Name":
            return t.get("Value", "") or ""
    return ""


def select_aws_profile() -> Optional[str]:
    """Interactively select an AWS profile."""
    session = boto3.session.Session()
    profiles = session.available_profiles
    if not profiles:
        print("No named AWS profiles found. Using default environment credentials (if any).")
        return None

    print("\nSelect an AWS CLI profile:")
    for i, p in enumerate(profiles, 1):
        print(f"  [{i}] {p}")
    while True:
        choice = input("Enter number (or press Enter for 'default credentials'): ").strip()
        if choice == "":
            return None
        if choice.isdigit() and 1 <= int(choice) <= len(profiles):
            return profiles[int(choice) - 1]
        print("Invalid selection. Try again.")


def session_for_profile(profile: Optional[str]) -> boto3.session.Session:
    """Create a boto3 session for the specified profile."""
    if profile is None:
        return boto3.session.Session()
    try:
        return boto3.session.Session(profile_name=profile)
    except ProfileNotFound as e:
        print(f"Profile not found: {e}. Falling back to default credentials.")
        return boto3.session.Session()


def select_region(sess: boto3.session.Session) -> str:
    """Interactively select an AWS region."""
    try:
        ec2 = sess.client("ec2", region_name="us-east-1")
        regions = ec2.describe_regions(AllRegions=True)["Regions"]
    except (BotoCoreError, ClientError):
        regions = [{"RegionName": r} for r in [
            "us-east-1", "us-east-2", "us-west-1", "us-west-2",
            "eu-west-1", "eu-west-2", "eu-central-1",
            "ap-south-1", "ap-southeast-1", "ap-southeast-2", "ap-northeast-1"
        ]]

    names = sorted([r["RegionName"] for r in regions])
    print("\nSelect an AWS region:")
    for i, r in enumerate(names, 1):
        print(f"  [{i}] {r}")

    default_index = None
    if sess.region_name and sess.region_name in names:
        default_index = names.index(sess.region_name) + 1

    while True:
        prompt = "Enter number"
        if default_index:
            prompt += f" (press Enter for [{default_index}] {names[default_index-1]})"
        prompt += ": "
        choice = input(prompt).strip()
        if choice == "":
            if default_index:
                return names[default_index - 1]
            else:
                print("Please pick a region.")
                continue
        if choice.isdigit() and 1 <= int(choice) <= len(names):
            return names[int(choice) - 1]
        print("Invalid selection. Try again.")


def generate_route_description(destination: str, target: str, origin: str = "") -> str:
    """Generate a descriptive explanation for a route."""
    if not destination or not target:
        return ""
    
    # Analyze destination
    dest_desc = ""
    if destination == "0.0.0.0/0":
        dest_desc = "Internet traffic"
    elif destination == "::/0":
        dest_desc = "IPv6 Internet traffic"
    elif destination.startswith("10."):
        dest_desc = f"Private network {destination}"
    elif destination.startswith("172.") and int(destination.split('.')[1]) >= 16 and int(destination.split('.')[1]) <= 31:
        dest_desc = f"Private network {destination}"
    elif destination.startswith("192.168."):
        dest_desc = f"Private network {destination}"
    elif destination.startswith("pl-"):
        dest_desc = f"Prefix list {destination}"
    else:
        dest_desc = f"Network {destination}"
    
    # Analyze target
    target_desc = ""
    if target == "local":
        target_desc = "local VPC"
    elif target.startswith("igw-"):
        target_desc = "Internet Gateway"
    elif target.startswith("nat-"):
        target_desc = "NAT Gateway"
    elif target.startswith("tgw-"):
        target_desc = "Transit Gateway"
    elif target.startswith("vgw-"):
        target_desc = "Virtual Private Gateway (VPN)"
    elif target.startswith("pcx-"):
        target_desc = "VPC Peering Connection"
    elif target.startswith("eni-"):
        target_desc = "Network Interface"
    elif target.startswith("i-"):
        target_desc = "EC2 Instance"
    elif target.startswith("eigw-"):
        target_desc = "Egress-only Internet Gateway"
    elif target.startswith("lgw-"):
        target_desc = "Local Gateway"
    elif target.startswith("cagw-"):
        target_desc = "Carrier Gateway"
    else:
        target_desc = f"target {target}"
    
    # Combine for description
    description = f"Route {dest_desc} to {target_desc}"
    
    # Add origin context if available
    if origin == "CreateRoute":
        description += " (custom route)"
    elif origin == "CreateRouteTable":
        description += " (default route)"
    
    return description


def get_vpc_route_tables_data(sess: boto3.session.Session, region: str) -> pd.DataFrame:
    """Get VPC route tables data and return as DataFrame."""
    ec2 = sess.client("ec2", region_name=region)

    try:
        # Get VPCs for context
        vpcs = ec2.describe_vpcs()["Vpcs"]
        vpc_lookup = {vpc["VpcId"]: {
            "name": tag_name(vpc.get("Tags")),
            "cidr": vpc.get("CidrBlock", "")
        } for vpc in vpcs}

        # Get subnets for association context
        subnets = ec2.describe_subnets()["Subnets"]
        subnet_lookup = {subnet["SubnetId"]: {
            "name": tag_name(subnet.get("Tags")),
            "cidr": subnet.get("CidrBlock", ""),
            "az": subnet.get("AvailabilityZone", "")
        } for subnet in subnets}

        # Get route tables
        route_tables = ec2.describe_route_tables()["RouteTables"]
        
        rows = []
        for rt in route_tables:
            rt_id = rt.get("RouteTableId", "")
            rt_name = tag_name(rt.get("Tags"))
            vpc_id = rt.get("VpcId", "")
            vpc_info = vpc_lookup.get(vpc_id, {"name": "", "cidr": ""})
            
            # Check if this is the main route table
            is_main = any(assoc.get("Main", False) for assoc in rt.get("Associations", []))
            
            # Get associated subnets
            associated_subnets = []
            for assoc in rt.get("Associations", []):
                if assoc.get("SubnetId"):
                    subnet_id = assoc["SubnetId"]
                    subnet_info = subnet_lookup.get(subnet_id, {"name": "", "cidr": "", "az": ""})
                    subnet_display = f"{subnet_info['name'] or subnet_id} ({subnet_info['cidr']}) - {subnet_info['az']}"
                    associated_subnets.append(subnet_display)
            
            associated_subnets_str = "\n".join(associated_subnets) if associated_subnets else ("Main Route Table" if is_main else "No associations")
            
            # Process routes
            routes = rt.get("Routes", [])
            if routes:
                for route in routes:
                    # Destination
                    destination = (route.get("DestinationCidrBlock") or 
                                 route.get("DestinationIpv6CidrBlock") or 
                                 route.get("DestinationPrefixListId") or "")
                    
                    # Target
                    target = (route.get("GatewayId") or 
                            route.get("NatGatewayId") or 
                            route.get("TransitGatewayId") or 
                            route.get("VpcPeeringConnectionId") or 
                            route.get("InstanceId") or 
                            route.get("NetworkInterfaceId") or 
                            route.get("EgressOnlyInternetGatewayId") or 
                            route.get("LocalGatewayId") or 
                            route.get("CarrierGatewayId") or "")
                    
                    state = route.get("State", "")
                    origin = route.get("Origin", "")
                    
                    # Generate description
                    description = generate_route_description(destination, target, origin)
                    
                    rows.append({
                        "Route Table Name": rt_name,
                        "Route Table ID": rt_id,
                        "VPC Name": vpc_info["name"],
                        "VPC ID": vpc_id,
                        "VPC CIDR": vpc_info["cidr"],
                        "Is Main": "Yes" if is_main else "No",
                        "Associated Subnets": associated_subnets_str,
                        "Destination": destination,
                        "Target": target,
                        "Description": description,
                        "State": state,
                        "Origin": origin
                    })
            else:
                # Route table with no routes
                rows.append({
                    "Route Table Name": rt_name,
                    "Route Table ID": rt_id,
                    "VPC Name": vpc_info["name"],
                    "VPC ID": vpc_id,
                    "VPC CIDR": vpc_info["cidr"],
                    "Is Main": "Yes" if is_main else "No",
                    "Associated Subnets": associated_subnets_str,
                    "Destination": "",
                    "Target": "",
                    "Description": "",
                    "State": "",
                    "Origin": ""
                })

        df = pd.DataFrame(rows, columns=[
            "Route Table Name", "Route Table ID", "VPC Name", "VPC ID", "VPC CIDR",
            "Is Main", "Associated Subnets", "Destination", "Target", "Description", "State", "Origin"
        ])
        
        return df

    except (BotoCoreError, ClientError) as e:
        print(f"Error fetching VPC route tables: {e}", file=sys.stderr)
        return pd.DataFrame()


def generate_tgw_route_description(destination: str, route_type: str, resource_type: str, resource_id: str) -> str:
    """Generate a descriptive explanation for a Transit Gateway route."""
    if not destination:
        return ""
    
    # Analyze destination
    if destination == "0.0.0.0/0":
        dest_desc = "Default route (all traffic)"
    elif destination.startswith("10."):
        dest_desc = f"Private network {destination}"
    elif destination.startswith("172.") and int(destination.split('.')[1]) >= 16 and int(destination.split('.')[1]) <= 31:
        dest_desc = f"Private network {destination}"
    elif destination.startswith("192.168."):
        dest_desc = f"Private network {destination}"
    else:
        dest_desc = f"Network {destination}"
    
    # Analyze target based on resource type
    if resource_type == "vpc":
        target_desc = f"VPC ({resource_id})"
    elif resource_type == "vpn":
        target_desc = f"VPN connection ({resource_id})"
    elif resource_type == "direct-connect-gateway":
        target_desc = f"Direct Connect Gateway ({resource_id})"
    elif resource_type == "peering":
        target_desc = f"Transit Gateway Peering ({resource_id})"
    elif resource_type == "connect":
        target_desc = f"Connect attachment ({resource_id})"
    elif resource_id:
        target_desc = f"{resource_type} ({resource_id})"
    else:
        target_desc = "unknown target"
    
    # Build description based on route type
    if route_type == "static":
        description = f"Static route: {dest_desc} → {target_desc}"
    elif route_type == "propagated":
        description = f"Propagated route: {dest_desc} → {target_desc}"
    else:
        description = f"Route: {dest_desc} → {target_desc}"
    
    return description


def get_tgw_route_tables_data(sess: boto3.session.Session, region: str) -> pd.DataFrame:
    """Get Transit Gateway route tables data and return as DataFrame."""
    ec2 = sess.client("ec2", region_name=region)

    try:
        # Get Transit Gateways
        tgws_response = ec2.describe_transit_gateways()
        tgws = tgws_response.get("TransitGateways", [])
        
        if not tgws:
            print("No Transit Gateways found in the region.")
            return pd.DataFrame()

        tgw_lookup = {tgw["TransitGatewayId"]: {
            "name": tag_name(tgw.get("Tags")),
            "state": tgw.get("State", ""),
            "asn": tgw.get("Options", {}).get("AmazonSideAsn", "")
        } for tgw in tgws}

        # Get Transit Gateway Route Tables
        tgw_rts_response = ec2.describe_transit_gateway_route_tables()
        tgw_route_tables = tgw_rts_response.get("TransitGatewayRouteTables", [])

        if not tgw_route_tables:
            print("No Transit Gateway Route Tables found.")
            return pd.DataFrame()

        rows = []
        
        for tgw_rt in tgw_route_tables:
            tgw_rt_id = tgw_rt.get("TransitGatewayRouteTableId", "")
            tgw_rt_name = tag_name(tgw_rt.get("Tags"))
            tgw_id = tgw_rt.get("TransitGatewayId", "")
            tgw_info = tgw_lookup.get(tgw_id, {"name": "", "state": "", "asn": ""})
            
            is_default = tgw_rt.get("DefaultAssociationRouteTable", False)
            is_propagation_default = tgw_rt.get("DefaultPropagationRouteTable", False)
            state = tgw_rt.get("State", "")

            try:
                # Get routes for this route table
                routes_response = ec2.search_transit_gateway_routes(
                    TransitGatewayRouteTableId=tgw_rt_id,
                    Filters=[{
                        "Name": "state",
                        "Values": ["active", "blackhole"]
                    }]
                )
                routes = routes_response.get("Routes", [])

                if routes:
                    for route in routes:
                        destination = route.get("DestinationCidrBlock", "")
                        state_route = route.get("State", "")
                        route_type = route.get("Type", "")
                        
                        # Get attachment information
                        attachments = route.get("TransitGatewayAttachments", [])
                        if attachments:
                            for attachment in attachments:
                                attachment_id = attachment.get("TransitGatewayAttachmentId", "")
                                resource_id = attachment.get("ResourceId", "")
                                resource_type = attachment.get("ResourceType", "")
                                
                                # Generate description
                                description = generate_tgw_route_description(destination, route_type, resource_type, resource_id)
                                
                                rows.append({
                                    "TGW Route Table Name": tgw_rt_name,
                                    "TGW Route Table ID": tgw_rt_id,
                                    "Transit Gateway Name": tgw_info["name"],
                                    "Transit Gateway ID": tgw_id,
                                    "TGW ASN": tgw_info["asn"],
                                    "TGW State": tgw_info["state"],
                                    "Is Default Association": "Yes" if is_default else "No",
                                    "Is Default Propagation": "Yes" if is_propagation_default else "No",
                                    "Route Table State": state,
                                    "Destination CIDR": destination,
                                    "Route State": state_route,
                                    "Route Type": route_type,
                                    "Description": description,
                                    "Attachment ID": attachment_id,
                                    "Resource ID": resource_id,
                                    "Resource Type": resource_type
                                })
                        else:
                            # Route without attachments
                            description = generate_tgw_route_description(destination, route_type, "", "")
                            
                            rows.append({
                                "TGW Route Table Name": tgw_rt_name,
                                "TGW Route Table ID": tgw_rt_id,
                                "Transit Gateway Name": tgw_info["name"],
                                "Transit Gateway ID": tgw_id,
                                "TGW ASN": tgw_info["asn"],
                                "TGW State": tgw_info["state"],
                                "Is Default Association": "Yes" if is_default else "No",
                                "Is Default Propagation": "Yes" if is_propagation_default else "No",
                                "Route Table State": state,
                                "Destination CIDR": destination,
                                "Route State": state_route,
                                "Route Type": route_type,
                                "Description": description,
                                "Attachment ID": "",
                                "Resource ID": "",
                                "Resource Type": ""
                            })
                else:
                    # Route table with no routes
                    rows.append({
                        "TGW Route Table Name": tgw_rt_name,
                        "TGW Route Table ID": tgw_rt_id,
                        "Transit Gateway Name": tgw_info["name"],
                        "Transit Gateway ID": tgw_id,
                        "TGW ASN": tgw_info["asn"],
                        "TGW State": tgw_info["state"],
                        "Is Default Association": "Yes" if is_default else "No",
                        "Is Default Propagation": "Yes" if is_propagation_default else "No",
                        "Route Table State": state,
                        "Destination CIDR": "",
                        "Route State": "",
                        "Route Type": "",
                        "Description": "",
                        "Attachment ID": "",
                        "Resource ID": "",
                        "Resource Type": ""
                    })

            except (BotoCoreError, ClientError) as e:
                print(f"Error fetching routes for TGW Route Table {tgw_rt_id}: {e}")
                # Add a row showing the route table even if we can't get its routes
                rows.append({
                    "TGW Route Table Name": tgw_rt_name,
                    "TGW Route Table ID": tgw_rt_id,
                    "Transit Gateway Name": tgw_info["name"],
                    "Transit Gateway ID": tgw_id,
                    "TGW ASN": tgw_info["asn"],
                    "TGW State": tgw_info["state"],
                    "Is Default Association": "Yes" if is_default else "No",
                    "Is Default Propagation": "Yes" if is_propagation_default else "No",
                    "Route Table State": state,
                    "Destination CIDR": "Error fetching routes",
                    "Route State": "",
                    "Route Type": "",
                    "Description": "Unable to fetch route details",
                    "Attachment ID": "",
                    "Resource ID": "",
                    "Resource Type": ""
                })

        df = pd.DataFrame(rows, columns=[
            "TGW Route Table Name", "TGW Route Table ID", "Transit Gateway Name", 
            "Transit Gateway ID", "TGW ASN", "TGW State", "Is Default Association", 
            "Is Default Propagation", "Route Table State", "Destination CIDR", 
            "Route State", "Route Type", "Description", "Attachment ID", "Resource ID", "Resource Type"
        ])
        
        return df

    except (BotoCoreError, ClientError) as e:
        print(f"Error fetching Transit Gateway route tables: {e}", file=sys.stderr)
        return pd.DataFrame()


def write_excel_with_sheets(vpc_df: pd.DataFrame, tgw_df: pd.DataFrame, outfile: str) -> None:
    """Write both DataFrames to separate sheets in Excel with formatting."""
    engine = "openpyxl"
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except Exception:
        pass

    with pd.ExcelWriter(outfile, engine=engine) as writer:
        # Write VPC Route Tables sheet
        if not vpc_df.empty:
            vpc_df.to_excel(writer, sheet_name="VPC Route Tables", index=False)
        
        # Write TGW Route Tables sheet
        if not tgw_df.empty:
            tgw_df.to_excel(writer, sheet_name="TGW Route Tables", index=False)

    # Add Excel Tables with styling via openpyxl
    wb = load_workbook(outfile)
    
    # Format VPC Route Tables sheet
    if "VPC Route Tables" in wb.sheetnames and not vpc_df.empty:
        ws_vpc = wb["VPC Route Tables"]
        max_row = ws_vpc.max_row
        max_col = ws_vpc.max_column
        if max_row > 1:  # Only create table if there's data
            last_col_letter = ws_vpc.cell(row=1, column=max_col).column_letter
            table_ref = f"A1:{last_col_letter}{max_row}"
            
            table = Table(displayName="VPCRouteTablesTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws_vpc.add_table(table)
            
            # Auto-fit column widths
            for col in ws_vpc.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                    except Exception:
                        val = ""
                    if len(val) > max_len:
                        max_len = len(val)
                ws_vpc.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    # Format TGW Route Tables sheet
    if "TGW Route Tables" in wb.sheetnames and not tgw_df.empty:
        ws_tgw = wb["TGW Route Tables"]
        max_row = ws_tgw.max_row
        max_col = ws_tgw.max_column
        if max_row > 1:  # Only create table if there's data
            last_col_letter = ws_tgw.cell(row=1, column=max_col).column_letter
            table_ref = f"A1:{last_col_letter}{max_row}"
            
            table = Table(displayName="TGWRouteTablesTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws_tgw.add_table(table)
            
            # Auto-fit column widths
            for col in ws_tgw.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                    except Exception:
                        val = ""
                    if len(val) > max_len:
                        max_len = len(val)
                ws_tgw.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    wb.save(outfile)


def main() -> int:
    print("=== AWS Route Tables -> Excel Export ===")
    profile = select_aws_profile()
    base_sess = session_for_profile(profile)
    region = select_region(base_sess)
    print(f"\nUsing profile: {profile or 'default'} | region: {region}")

    print("\nFetching VPC Route Tables...")
    vpc_df = get_vpc_route_tables_data(base_sess, region)
    
    print("Fetching Transit Gateway Route Tables...")
    tgw_df = get_tgw_route_tables_data(base_sess, region)

    if vpc_df.empty and tgw_df.empty:
        print("No route table data found.")
        return 0

    # Output filename
    safe_region = region.replace(":", "_")
    outfile = f"aws_route_tables_export_{safe_region}.xlsx"
    
    write_excel_with_sheets(vpc_df, tgw_df, outfile)
    
    print(f"\nExport complete: {outfile}")
    if not vpc_df.empty:
        print(f"  - VPC Route Tables: {len(vpc_df)} entries")
    if not tgw_df.empty:
        print(f"  - TGW Route Tables: {len(tgw_df)} entries")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())