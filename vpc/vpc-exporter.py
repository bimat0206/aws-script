#!/usr/bin/env python3
"""
Export AWS VPC information (VPCs, Subnets, Route Tables) to an Excel file.

Features
- Interactive AWS profile and region selection
- Exports VPC name, CIDR; Subnet names & CIDRs; Route Table names & detailed routes
- Generates .xlsx with a structured, formatted table (single sheet: "VPCs")

Requirements:
    pip install boto3 pandas openpyxl
Optional (faster Excel writing):
    pip install xlsxwriter

Usage:
    python export_vpc_to_xlsx.py
"""

import sys
import os
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from collections import defaultdict

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError, ProfileNotFound
except ImportError:
    print("Missing dependency: boto3. Install with `pip install boto3`.", file=sys.stderr)
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("Missing dependency: pandas. Install with `pip install pandas`.", file=sys.stderr)
    sys.exit(1)

# openpyxl is used to create a real 'Table' object in the Excel sheet
try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Missing dependency: openpyxl. Install with `pip install openpyxl`.", file=sys.stderr)
    sys.exit(1)


@dataclass
class RouteInfo:
    destination: str
    target: str
    state: str


def tag_name(tags: Optional[List[Dict]]) -> str:
    if not tags:
        return ""
    for t in tags:
        if t.get("Key") == "Name":
            return t.get("Value", "") or ""
    return ""


def select_aws_profile() -> Optional[str]:
    session = boto3.session.Session()
    profiles = session.available_profiles
    if not profiles:
        print("No named AWS profiles found. Using default environment credentials (if any).")
        return None

    print("\nSelect an AWS CLI profile:")
    for i, p in enumerate(profiles, 1):
        print(f"  [{i}] {p}")
    while True:
        choice = input("Enter number (or press Enter for 'default credentials'): ").strip()
        if choice == "":
            return None
        if choice.isdigit() and 1 <= int(choice) <= len(profiles):
            return profiles[int(choice) - 1]
        print("Invalid selection. Try again.")


def session_for_profile(profile: Optional[str]) -> boto3.session.Session:
    if profile is None:
        return boto3.session.Session()
    try:
        return boto3.session.Session(profile_name=profile)
    except ProfileNotFound as e:
        print(f"Profile not found: {e}. Falling back to default credentials.")
        return boto3.session.Session()


def select_region(sess: boto3.session.Session) -> str:
    # Use a generic region to list all regions if no default is configured
    try:
        ec2 = sess.client("ec2", region_name="us-east-1")
        regions = ec2.describe_regions(AllRegions=True)["Regions"]
    except (BotoCoreError, ClientError):
        # Fallback to a default set if describe_regions fails
        regions = [{"RegionName": r} for r in [
            "us-east-1", "us-east-2", "us-west-1", "us-west-2",
            "eu-west-1", "eu-west-2", "eu-central-1",
            "ap-south-1", "ap-southeast-1", "ap-southeast-2", "ap-northeast-1"
        ]]

    names = sorted([r["RegionName"] for r in regions])
    print("\nSelect an AWS region:")
    for i, r in enumerate(names, 1):
        print(f"  [{i}] {r}")

    default_index = None
    if sess.region_name and sess.region_name in names:
        default_index = names.index(sess.region_name) + 1

    while True:
        prompt = "Enter number"
        if default_index:
            prompt += f" (press Enter for [{default_index}] {names[default_index-1]})"
        prompt += ": "
        choice = input(prompt).strip()
        if choice == "":
            if default_index:
                return names[default_index - 1]
            else:
                print("Please pick a region.")
                continue
        if choice.isdigit() and 1 <= int(choice) <= len(names):
            return names[int(choice) - 1]
        print("Invalid selection. Try again.")


def get_vpc_data(sess: boto3.session.Session, region: str) -> pd.DataFrame:
    ec2 = sess.client("ec2", region_name=region)

    # Describe VPCs, Subnets, Route Tables
    vpcs = ec2.describe_vpcs()["Vpcs"]
    subnets = ec2.describe_subnets()["Subnets"]
    rts = ec2.describe_route_tables()["RouteTables"]

    # Build helpers
    vpc_id_to_main_rt: Dict[str, Dict] = {}
    subnet_to_rt: Dict[str, Dict] = {}

    for rt in rts:
        associations = rt.get("Associations", [])
        for assoc in associations:
            if assoc.get("Main"):
                vpc_id = rt.get("VpcId")
                if vpc_id:
                    vpc_id_to_main_rt[vpc_id] = rt
            if assoc.get("SubnetId"):
                subnet_to_rt[assoc["SubnetId"]] = rt

    # Prepare rows
    rows: List[Dict] = []

    # Group subnets by VPC
    subnets_by_vpc: Dict[str, List[Dict]] = defaultdict(list)
    for sn in subnets:
        subnets_by_vpc[sn["VpcId"]].append(sn)

    for vpc in vpcs:
        vpc_id = vpc["VpcId"]
        vpc_cidrs = [cidr["CidrBlock"] for cidr in vpc.get("CidrBlockAssociationSet", []) if cidr.get("CidrBlock")]
        if not vpc_cidrs and vpc.get("CidrBlock"):
            vpc_cidrs = [vpc["CidrBlock"]]
        vpc_cidr = ", ".join(vpc_cidrs)
        vpc_name = tag_name(vpc.get("Tags"))
        vpc_row_base = {
            "VPC Name": vpc_name,
            "VPC ID": vpc_id,
            "VPC CIDR": vpc_cidr,
        }

        vpc_subnets = subnets_by_vpc.get(vpc_id, [])
        if not vpc_subnets:
            # still produce a row with empty subnet/route table
            rows.append({**vpc_row_base,
                         "Subnet Name": "", "Subnet ID": "", "Subnet CIDR": "",
                         "Route Table Name": "", "Route Table ID": "",
                         "Route Destinations": "", "Route Targets": "", "Route States": ""})
            continue

        for sn in sorted(vpc_subnets, key=lambda s: tag_name(s.get("Tags")) or s["SubnetId"]):
            subnet_id = sn["SubnetId"]
            subnet_cidr = sn.get("CidrBlock", "")
            subnet_name = tag_name(sn.get("Tags"))

            rt = subnet_to_rt.get(subnet_id)
            if not rt:
                rt = vpc_id_to_main_rt.get(vpc_id)

            if rt:
                rt_name = tag_name(rt.get("Tags"))
                rt_id = rt.get("RouteTableId", "")
                # Extract routing details
                dsts, tgts, states = [], [], []
                for r in rt.get("Routes", []):
                    dst = r.get("DestinationCidrBlock") or r.get("DestinationIpv6CidrBlock") or r.get("DestinationPrefixListId") or ""
                    target = r.get("GatewayId") or r.get("NatGatewayId") or r.get("TransitGatewayId") or r.get("VpcPeeringConnectionId") or r.get("InstanceId") or r.get("NetworkInterfaceId") or r.get("EgressOnlyInternetGatewayId") or ""
                    state = r.get("State", "")
                    dsts.append(dst)
                    tgts.append(target)
                    states.append(state)
                route_dests = "\n".join(dsts)
                route_tgts = "\n".join(tgts)
                route_states = "\n".join(states)
            else:
                rt_name = ""
                rt_id = ""
                route_dests = route_tgts = route_states = ""

            rows.append({
                **vpc_row_base,
                "Subnet Name": subnet_name,
                "Subnet ID": subnet_id,
                "Subnet CIDR": subnet_cidr,
                "Route Table Name": rt_name,
                "Route Table ID": rt_id,
                "Route Destinations": route_dests,
                "Route Targets": route_tgts,
                "Route States": route_states,
            })

    df = pd.DataFrame(rows, columns=[
        "VPC Name", "VPC ID", "VPC CIDR",
        "Subnet Name", "Subnet ID", "Subnet CIDR",
        "Route Table Name", "Route Table ID",
        "Route Destinations", "Route Targets", "Route States"
    ])
    return df


def write_excel(df: pd.DataFrame, outfile: str) -> None:
    # Write with pandas first
    engine = "openpyxl"
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except Exception:
        pass

    with pd.ExcelWriter(outfile, engine=engine) as writer:
        sheet_name = "VPCs"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()

    # Add an Excel Table with styling via openpyxl regardless of engine
    wb = load_workbook(outfile)
    ws = wb["VPCs"]
    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = ws.cell(row=1, column=max_col).column_letter
    table_ref = f"A1:{last_col_letter}{max_row}"

    table = Table(displayName="VPCsTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-fit-ish column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    wb.save(outfile)


def main() -> int:
    print("=== AWS VPC -> Excel Export ===")
    profile = select_aws_profile()
    base_sess = session_for_profile(profile)
    region = select_region(base_sess)
    print(f"\nUsing profile: {profile or 'default'} | region: {region}")

    try:
        df = get_vpc_data(base_sess, region)
    except (BotoCoreError, ClientError) as e:
        print(f"Error fetching data from AWS: {e}", file=sys.stderr)
        return 2

    if df.empty:
        print("No VPC data found.")
        return 0

    # Output filename
    safe_region = region.replace(":", "_")
    outfile = f"aws_vpc_export_{safe_region}.xlsx"
    write_excel(df, outfile)
    print(f"\nExport complete: {outfile}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
